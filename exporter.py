import csv
from collections import Counter
from pathlib import Path
from typing import List

from openpyxl import Workbook
from models import Game


def export_to_csv(games: List[Game], output_path: str = "output/library.csv"):
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    headers = [
        "AppID", "Name", "Playtime (Hours)", "Last Played", 
        "Release Date", "Developers", "Publishers", 
        "Price", "Discount (%)"
    ]
    
    with open(output_path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        
        for game in games:
            devs = ", ".join(game.developers) if game.developers else ""
            pubs = ", ".join(game.publishers) if game.publishers else ""
            last_played = game.last_played.strftime("%Y-%m-%d %H:%M") if game.last_played else "Never"
            
            writer.writerow([
                game.appid,
                game.name,
                round(game.playtime_hours, 1),
                last_played,
                game.release_date or "",
                devs,
                pubs,
                game.price or "",
                game.discount if game.discount is not None else ""
            ])


def export_to_excel(games: List[Game], output_path: str = "output/library.xlsx"):
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    
    # --- Sheet 1: Library ---
    ws_library = wb.active
    ws_library.title = "Library"
    
    headers = [
        "AppID", "Name", "Playtime (Hours)", "Last Played", 
        "Release Date", "Developers", "Publishers", 
        "Price", "Discount (%)"
    ]
    ws_library.append(headers)
    
    for game in games:
        devs = ", ".join(game.developers) if game.developers else ""
        pubs = ", ".join(game.publishers) if game.publishers else ""
        last_played = game.last_played.strftime("%Y-%m-%d %H:%M") if game.last_played else "Never"
        
        ws_library.append([
            game.appid,
            game.name,
            round(game.playtime_hours, 1),
            last_played,
            game.release_date or "",
            devs,
            pubs,
            game.price or "",
            game.discount if game.discount is not None else ""
        ])

    # --- Sheet 2: Statistics ---
    ws_stats = wb.create_sheet(title="Statistics")
    
    total_playtime = sum(g.playtime_hours for g in games)
    paid_games = sum(1 for g in games if not g.is_free)
    free_games = len(games) - paid_games
    
    ws_stats.append(["Metric", "Value"])
    ws_stats.append(["Total Games", len(games)])
    ws_stats.append(["Total Playtime (Hours)", round(total_playtime, 1)])
    ws_stats.append(["Paid Games", paid_games])
    ws_stats.append(["Free Games", free_games])

    # --- Sheet 3: Developers ---
    ws_devs = wb.create_sheet(title="Developers")
    dev_counts = Counter(d for game in games for d in game.developers)
    ws_devs.append(["Developer", "Game Count"])
    for dev, count in dev_counts.most_common():
        ws_devs.append([dev, count])

    # --- Sheet 4: Publishers ---
    ws_pubs = wb.create_sheet(title="Publishers")
    pub_counts = Counter(p for game in games for p in game.publishers)
    ws_pubs.append(["Publisher", "Game Count"])
    for pub, count in pub_counts.most_common():
        ws_pubs.append([pub, count])

    wb.save(output_path)